"""
Storage Adapter Layer for SheetManipulator.

Provides a format-agnostic abstraction (ABC) over physical .xlsx and .csv files.
All write operations are protected by the FileLock mechanism from Task 02.
"""
import abc
import csv
import os
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from backend.core.locking import FileLock


# ---------------------------------------------------------------------------
# Type coercion helpers
# ---------------------------------------------------------------------------

_TYPE_COERCIONS = {
    "int": int,
    "float": float,
    "boolean": lambda v: str(v).strip().lower() in ("true", "1", "yes"),
    "date": lambda v: datetime.strptime(str(v), "%Y-%m-%d").date() if v and str(v).strip() else None,
    "string": str,
}


def _coerce_value(value: Any, field_type: str) -> Any:
    """
    Convert a raw string/cell value to the target Python type.
    Returns None for empty/null values.
    """
    if value is None or str(value).strip() == "" or str(value).strip().lower() == "none":
        return None

    coerce_fn = _TYPE_COERCIONS.get(field_type, str)
    try:
        return coerce_fn(value)
    except (ValueError, TypeError):
        return value  # Return as-is on coercion failure


def _coerce_row(row: Dict[str, Any], field_map: Dict[str, str]) -> Dict[str, Any]:
    """Apply type coercion to every field in a row according to field_map."""
    return {
        key: _coerce_value(val, field_map.get(key, "string"))
        for key, val in row.items()
    }


def _build_field_map(fields_config: List[Dict[str, Any]]) -> Dict[str, str]:
    """Build {field_name: field_type} mapping from config fields list."""
    return {f["name"]: f.get("type", "string") for f in fields_config}


# ---------------------------------------------------------------------------
# Abstract Base Class
# ---------------------------------------------------------------------------

class BaseStorageAdapter(abc.ABC):
    """
    Abstract storage adapter defining the contract for all file format implementations.

    All adapters are initialised with the entity's storage config dict and
    an ordered list of field config dicts for type coercion.
    """

    def __init__(
        self,
        file_path: str,
        storage_config: Dict[str, Any],
        fields_config: List[Dict[str, Any]],
        lock_timeout: float = 5.0,
    ):
        self.file_path = os.path.abspath(file_path)
        self.storage_config = storage_config
        self.fields_config = fields_config
        self.field_map = _build_field_map(fields_config)
        self.field_names = [f["name"] for f in fields_config]
        self.lock_timeout = lock_timeout

    @abc.abstractmethod
    def read_all(self) -> List[Dict[str, Any]]:
        """
        Read all rows from the storage file.
        Returns a list of dicts where keys are column names.
        """
        ...

    @abc.abstractmethod
    def write_all(self, data: List[Dict[str, Any]]) -> None:
        """
        Overwrite the entire storage file with the provided data.
        Must be wrapped with FileLock.
        """
        ...

    @abc.abstractmethod
    def append_row(self, row: Dict[str, Any]) -> None:
        """
        Append a single row to the storage file.
        Must be wrapped with FileLock.
        """
        ...


# ---------------------------------------------------------------------------
# CSV Adapter
# ---------------------------------------------------------------------------

class CSVAdapter(BaseStorageAdapter):
    """
    Storage adapter for .csv files using Python's built-in csv module.
    Respects `delimiter` and `encoding` from the storage settings.
    """

    def __init__(self, file_path: str, storage_config: Dict[str, Any],
                 fields_config: List[Dict[str, Any]], lock_timeout: float = 5.0):
        super().__init__(file_path, storage_config, fields_config, lock_timeout)
        settings = storage_config.get("settings", {})
        self.delimiter = settings.get("delimiter", ",")
        self.encoding = settings.get("encoding", "utf-8")

    def read_all(self) -> List[Dict[str, Any]]:
        """Read all rows from the CSV file with type coercion applied."""
        if not os.path.isfile(self.file_path):
            return []

        rows = []
        with open(self.file_path, newline="", encoding=self.encoding) as f:
            reader = csv.DictReader(f, delimiter=self.delimiter)
            for raw_row in reader:
                rows.append(_coerce_row(dict(raw_row), self.field_map))
        return rows

    def write_all(self, data: List[Dict[str, Any]]) -> None:
        """Overwrite the CSV file with the provided data (FileLock protected)."""
        # Ensure parent directory exists
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

        with FileLock(self.file_path, timeout=self.lock_timeout):
            with open(self.file_path, "w", newline="", encoding=self.encoding) as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=self.field_names,
                    delimiter=self.delimiter,
                    extrasaction="ignore",
                )
                writer.writeheader()
                for row in data:
                    writer.writerow({k: ("" if row.get(k) is None else row[k]) for k in self.field_names})

    def append_row(self, row: Dict[str, Any]) -> None:
        """Append a single row to the CSV file (FileLock protected)."""
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        file_exists = os.path.isfile(self.file_path)

        with FileLock(self.file_path, timeout=self.lock_timeout):
            with open(self.file_path, "a", newline="", encoding=self.encoding) as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=self.field_names,
                    delimiter=self.delimiter,
                    extrasaction="ignore",
                )
                if not file_exists:
                    writer.writeheader()
                writer.writerow({k: ("" if row.get(k) is None else row[k]) for k in self.field_names})


# ---------------------------------------------------------------------------
# Excel Adapter
# ---------------------------------------------------------------------------

class ExcelAdapter(BaseStorageAdapter):
    """
    Storage adapter for .xlsx files using openpyxl.
    Respects `sheet_name` and `header_row` from the storage settings.
    """

    def __init__(self, file_path: str, storage_config: Dict[str, Any],
                 fields_config: List[Dict[str, Any]], lock_timeout: float = 5.0):
        super().__init__(file_path, storage_config, fields_config, lock_timeout)
        settings = storage_config.get("settings", {})
        self.sheet_name = settings.get("sheet_name", None)  # None = active sheet

    def _get_openpyxl(self):
        try:
            import openpyxl  # type: ignore
            return openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for ExcelAdapter. "
                "Install with: pip install openpyxl"
            ) from exc

    def read_all(self) -> List[Dict[str, Any]]:
        """Read all rows from the Excel file with type coercion applied."""
        if not os.path.isfile(self.file_path):
            return []

        openpyxl = self._get_openpyxl()
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)

        if self.sheet_name and self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            return []

        headers = [str(cell) if cell is not None else "" for cell in all_rows[0]]
        results = []
        for raw_row in all_rows[1:]:
            row_dict = {headers[i]: raw_row[i] for i in range(min(len(headers), len(raw_row)))}
            results.append(_coerce_row(row_dict, self.field_map))

        return results

    def write_all(self, data: List[Dict[str, Any]]) -> None:
        """Overwrite the Excel file with the provided data (FileLock protected)."""
        openpyxl = self._get_openpyxl()
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

        with FileLock(self.file_path, timeout=self.lock_timeout):
            wb = openpyxl.Workbook()
            ws = wb.active
            if self.sheet_name:
                ws.title = self.sheet_name

            # Write header
            ws.append(self.field_names)

            # Write data rows
            for row in data:
                ws.append([row.get(k) for k in self.field_names])

            wb.save(self.file_path)

    def append_row(self, row: Dict[str, Any]) -> None:
        """Append a single row to the Excel file (FileLock protected)."""
        openpyxl = self._get_openpyxl()
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

        with FileLock(self.file_path, timeout=self.lock_timeout):
            if os.path.isfile(self.file_path):
                wb = openpyxl.load_workbook(self.file_path)
                if self.sheet_name and self.sheet_name in wb.sheetnames:
                    ws = wb[self.sheet_name]
                else:
                    ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                if self.sheet_name:
                    ws.title = self.sheet_name
                # Write header for new file
                ws.append(self.field_names)

            ws.append([row.get(k) for k in self.field_names])
            wb.save(self.file_path)


# ---------------------------------------------------------------------------
# Adapter Factory
# ---------------------------------------------------------------------------

def get_adapter_for_entity(
    entity_config: Dict[str, Any],
    base_dir: str = ".",
    lock_timeout: float = 5.0,
) -> BaseStorageAdapter:
    """
    Return the correct adapter instance based on the entity's format config.

    Args:
        entity_config: A single entity dict from config["entities"].
        base_dir: Base directory for resolving relative file_path values.
        lock_timeout: Timeout for FileLock acquisition (seconds).

    Returns:
        A `CSVAdapter` or `ExcelAdapter` instance.

    Raises:
        ValueError: If the format is not supported or config is malformed.
    """
    storage = entity_config.get("storage", {})
    fields = entity_config.get("fields", [])
    entity_name = entity_config.get("name", "<unnamed>")

    relative_path = storage.get("file_path", "")
    abs_path = os.path.join(os.path.abspath(base_dir), relative_path)
    fmt = storage.get("format", "").lower()

    if fmt == "csv":
        return CSVAdapter(abs_path, storage, fields, lock_timeout=lock_timeout)
    elif fmt == "xlsx":
        return ExcelAdapter(abs_path, storage, fields, lock_timeout=lock_timeout)
    else:
        raise ValueError(
            f"Entity '{entity_name}': unsupported storage format '{fmt}'. "
            "Must be 'csv' or 'xlsx'."
        )
