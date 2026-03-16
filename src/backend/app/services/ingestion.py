"""
Background ingestion pipeline.

Flow:
  1. Load workbook from cold storage.
  2. Fetch schema_definitions for the target table.
  3. Run validation engine (short-circuits at 1,000 errors).
  4a. On errors  → generate error report → update upload_registry status="error".
  4b. On success → ensure dynamic table exists → bulk-insert rows → status="ingested".

Note: dynamic_ddl.ensure_table_exists() is implemented in T-004.
      A lightweight stub is called here so the pipeline is wired end-to-end.
"""
from __future__ import annotations

import json
import os
import uuid
from datetime import datetime, timezone

import openpyxl
from sqlalchemy import select, text

from app.database import AsyncSessionLocal
from app.models.schema_definition import SchemaDefinition
from app.models.upload_registry import UploadRegistry
from app.services.export import generate_error_report
from app.services.validation import validate_worksheet


async def run_ingestion_pipeline(
    upload_id: str,
    worksheet_name: str,
    table_system_name: str,
) -> None:
    """
    Entry point for the background ingestion task.
    Called via FastAPI BackgroundTasks — must not raise unhandled exceptions.
    """
    async with AsyncSessionLocal() as session:
        try:
            # ── 1. Load workbook ──────────────────────────────────────────────
            workbook_path = f"/data/uploads/{upload_id}/original.xlsx"
            wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

            if worksheet_name not in wb.sheetnames:
                await _set_status(
                    session,
                    upload_id,
                    "error",
                    error_report_path=None,
                    error_detail=f"Worksheet '{worksheet_name}' not found in workbook.",
                )
                return

            ws = wb[worksheet_name]

            # ── 2. Fetch schema definitions ───────────────────────────────────
            result = await session.execute(
                select(SchemaDefinition)
                .where(SchemaDefinition.table_system_name == table_system_name)
                .order_by(SchemaDefinition.column_order)
            )
            schema: list[SchemaDefinition] = list(result.scalars().all())

            # ── 3. Validate ───────────────────────────────────────────────────
            errors = validate_worksheet(ws, schema)
            wb.close()

            # ── 4a. Validation failed ─────────────────────────────────────────
            if errors:
                report_path = generate_error_report(errors, upload_id)
                await _set_status(
                    session,
                    upload_id,
                    "error",
                    error_report_path=report_path,
                )
                return

            # ── 4b. Validation passed — bulk insert ───────────────────────────
            # Ensure the dynamic table exists (stub until T-004 is implemented).
            await _ensure_table_exists(session, table_system_name, schema)

            # Re-open workbook for data extraction (read_only iterates once)
            wb2 = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
            ws2 = wb2[worksheet_name]

            rows_iter = ws2.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                wb2.close()
                await _set_status(session, upload_id, "ingested")
                return

            headers = list(header_row)
            col_names = [col.column_system_name for col in schema]

            insert_rows: list[dict] = []
            for row in rows_iter:
                row_dict: dict = {
                    "_row_id": str(uuid.uuid4()),
                    "is_deleted": False,
                    "_upload_id": upload_id,
                }
                for col_def in schema:
                    try:
                        idx = headers.index(col_def.column_system_name)
                        row_dict[col_def.column_system_name] = (
                            row[idx] if idx < len(row) else None
                        )
                    except ValueError:
                        row_dict[col_def.column_system_name] = None
                insert_rows.append(row_dict)

            wb2.close()

            if insert_rows:
                # Build parameterized INSERT using SQLAlchemy text()
                # Column list is derived from schema — safe, not user-supplied raw SQL
                all_cols = ["_row_id", "is_deleted", "_upload_id"] + col_names
                col_clause = ", ".join(f'"{c}"' for c in all_cols)
                param_clause = ", ".join(f":{c}" for c in all_cols)
                stmt = text(
                    f'INSERT INTO "{table_system_name}" ({col_clause}) '
                    f"VALUES ({param_clause})"
                )
                await session.execute(stmt, insert_rows)

            await _set_status(session, upload_id, "ingested")

        except Exception as exc:  # noqa: BLE001
            # Catch-all: mark upload as error so the UI is never stuck on "validating"
            try:
                await _set_status(session, upload_id, "error")
            except Exception:  # noqa: BLE001
                pass
            raise exc


# ── Helpers ───────────────────────────────────────────────────────────────────

async def _set_status(
    session,
    upload_id: str,
    status: str,
    error_report_path: str | None = None,
    error_detail: str | None = None,
) -> None:
    """Update upload_registry status (and optionally error_report_path)."""
    result = await session.execute(
        select(UploadRegistry).where(UploadRegistry.upload_id == upload_id)
    )
    record: UploadRegistry | None = result.scalar_one_or_none()
    if record:
        record.status = status
        if error_report_path is not None:
            record.error_report_path = error_report_path
        await session.commit()


async def _ensure_table_exists(session, table_system_name: str, schema: list) -> None:
    """
    Lightweight DDL stub — creates the dynamic table if it does not exist.
    Full implementation lives in T-004 (services/dynamic_ddl.py).
    This stub is intentionally minimal so T-003 is self-contained.
    """
    # Map schema data_type → SQLite column type
    _type_map = {
        "String": "TEXT",
        "Integer": "INTEGER",
        "Float": "REAL",
        "Boolean": "INTEGER",  # SQLite stores booleans as 0/1
        "Date": "TEXT",
    }

    col_defs = [
        '"_row_id" TEXT PRIMARY KEY',
        '"is_deleted" INTEGER NOT NULL DEFAULT 0',
        '"_upload_id" TEXT',
    ]
    for col in schema:
        sql_type = _type_map.get(col.data_type, "TEXT")
        nullable = "" if col.is_mandatory else ""
        col_defs.append(f'"{col.column_system_name}" {sql_type}')

    ddl = (
        f'CREATE TABLE IF NOT EXISTS "{table_system_name}" '
        f"({', '.join(col_defs)})"
    )
    await session.execute(text(ddl))
    await session.commit()
